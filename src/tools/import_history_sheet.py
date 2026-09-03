from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
from typing import Any

from openpyxl import load_workbook


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def coerce_int(value: Any) -> int | str:
    value = clean_value(value)
    if value == "":
      return ""
    if isinstance(value, bool):
      return int(value)
    if isinstance(value, (int, float)):
      return int(value)
    text = str(value).strip()
    return int(float(text)) if text else ""


def coerce_float(value: Any) -> float | int | str:
    value = clean_value(value)
    if value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return float(text) if text else ""


def coerce_bool(value: Any) -> bool:
    value = clean_value(value)
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "yes", "1", "y"}


def split_list(value: Any) -> list[str]:
    raw = str(clean_value(value))
    if not raw:
        return []
    parts = re.split(r"\s*[|;]\s*|\s*,\s*", raw)
    return [part for part in (piece.strip() for piece in parts) if part]


def load_sheet_rows(workbook, name: str) -> list[dict[str, Any]]:
    if name not in workbook.sheetnames:
        raise ValueError(f"Missing required sheet '{name}' in workbook.")

    sheet = workbook[name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]
    data = []
    for raw_row in rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if header:
                row[header] = clean_value(raw_row[index] if index < len(raw_row) else "")
        if any(value != "" for value in row.values()) and not row_looks_like_header(row, headers):
            data.append(row)
    return data


def row_looks_like_header(row: dict[str, Any], headers: list[str]) -> bool:
    comparable_headers = [header for header in headers if header]
    if not comparable_headers:
        return False

    matching_cells = 0
    populated_cells = 0
    for header in comparable_headers:
      value = normalize_header(row.get(header, ""))
      if value:
          populated_cells += 1
          if value == header:
              matching_cells += 1

    return populated_cells > 0 and matching_cells >= max(2, populated_cells // 2)


def pick(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        normalized = normalize_header(key)
        if normalized in row and row[normalized] != "":
            return row[normalized]
    return default


def parse_league(workbook) -> dict[str, Any]:
    rows = load_sheet_rows(workbook, "LEAGUE")
    if not rows:
        raise ValueError("Sheet 'LEAGUE' is empty.")

    row = rows[0]
    return {
        "abbreviation": pick(row, "abbreviation", "abbr"),
        "fullName": pick(row, "fullName", "full name", "leagueName", "league name"),
        "started": coerce_int(pick(row, "started", "firstYear", "first year")),
        "currentYear": coerce_int(pick(row, "currentYear", "current year")),
        "currentDate": pick(row, "currentDate", "current date"),
        "commissioner": pick(row, "commissioner"),
    }


def parse_history(workbook) -> list[dict[str, Any]]:
    rows = load_sheet_rows(workbook, "HISTORY")
    items = []
    for row in rows:
        year = coerce_int(pick(row, "year"))
        if year == "":
            continue
        items.append(
            {
                "year": year,
                "champion": pick(row, "champion"),
                "northChampion": pick(row, "northChampion", "north champion"),
                "southChampion": pick(row, "southChampion", "south champion"),
                "championshipMvp": pick(row, "championshipMvp", "championship mvp"),
                "bestStandingsTeam": pick(row, "bestStandingsTeam", "best standings team"),
                "bestStandingsRecord": pick(row, "bestStandingsRecord", "best standings record"),
                "worstStandingsTeam": pick(row, "worstStandingsTeam", "worst standings team"),
                "worstStandingsRecord": pick(row, "worstStandingsRecord", "worst standings record"),
                "northMvp": pick(row, "northMvp", "north mvp"),
                "southMvp": pick(row, "southMvp", "south mvp"),
                "topSigning": pick(row, "topSigning", "top signing"),
                "topTrade": pick(row, "topTrade", "top trade"),
                "prospectsGameMvp": pick(row, "prospectsGameMvp", "prospects game mvp"),
            }
        )
    return items


def parse_teams(workbook) -> list[dict[str, Any]]:
    rows = load_sheet_rows(workbook, "TEAMS")
    items = []
    for row in rows:
        full_name = pick(row, "fullName", "full name", "team", "team name")
        if not full_name:
            continue
        items.append(
            {
                "fullName": full_name,
                "firstYear": coerce_int(pick(row, "firstYear", "first year")),
                "conference": pick(row, "conference"),
                "division": pick(row, "division"),
                "championships": coerce_int(pick(row, "championships")),
                "conferenceTitles": coerce_int(pick(row, "conferenceTitles", "conference titles")),
                "divisionTitles": coerce_int(pick(row, "divisionTitles", "division titles")),
                "playoffAppearances": coerce_int(pick(row, "playoffAppearances", "playoff appearances")),
                "manager": pick(row, "manager"),
                "rivals": split_list(pick(row, "rivals")),
                "traits": split_list(pick(row, "traits")),
                "biggestSigning": pick(row, "biggestSigning", "biggest signing"),
            }
        )
    return items


def parse_greats(workbook) -> list[dict[str, Any]]:
    rows = load_sheet_rows(workbook, "GREATS")
    items = []
    for row in rows:
        name = pick(row, "name", "player")
        if not name:
            continue
        items.append(
            {
                "name": name,
                "active": coerce_bool(pick(row, "active")),
                "seasons": coerce_int(pick(row, "seasons")),
                "mvps": coerce_int(pick(row, "mvps")),
                "allStars": coerce_int(pick(row, "allStars", "all stars")),
                "championships": coerce_int(pick(row, "championships")),
                "playoffMvps": coerce_int(pick(row, "playoffMvps", "playoff mvps")),
                "platinumSticks": coerce_int(pick(row, "platinumSticks", "platinum sticks")),
                "battingTitles": coerce_int(pick(row, "battingTitles", "batting titles")),
                "homeRuns": coerce_int(pick(row, "homeRuns", "home runs")),
                "war": coerce_float(pick(row, "war")),
                "hits": coerce_int(pick(row, "hits")),
                "opsPlus": coerce_int(pick(row, "opsPlus", "ops+")),
                "notes": pick(row, "notes"),
            }
        )
    return items


def parse_records(workbook) -> dict[str, list[dict[str, Any]]]:
    rows = load_sheet_rows(workbook, "RECORDS")
    records = {"career": [], "season": []}
    for row in rows:
        category = pick(row, "category")
        holder = pick(row, "holder", "player")
        value = pick(row, "value")
        if not category or not holder or value == "":
            continue

        year = coerce_int(pick(row, "year"))
        scope = str(pick(row, "scope", "type", "recordType", "record type")).strip().lower()
        if scope not in {"career", "season"}:
            scope = "season" if year != "" else "career"

        item = {
            "category": category,
            "holder": holder,
            "value": str(value),
        }
        if scope == "season" and year != "":
            item["year"] = year
        records[scope].append(item)
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="Import a Google Sheets Excel export into historySheet.mjs.")
    parser.add_argument("--input", required=True, help="Path to the .xlsx export file.")
    parser.add_argument("--output", required=True, help="Path to write the historySheet.mjs file.")
    parser.add_argument("--title", default="OOTP2", help="Source title to store in the history snapshot.")
    parser.add_argument("--spreadsheet-id", default="", help="Spreadsheet ID to store in the history snapshot.")
    parser.add_argument("--url", default="", help="Spreadsheet URL to store in the history snapshot.")
    args = parser.parse_args()

    workbook = load_workbook(args.input, data_only=True)
    synced_on = dt.date.today().isoformat()

    payload = {
        "source": {
            "title": args.title,
            "spreadsheetId": args.spreadsheet_id,
            "url": args.url,
            "syncedOn": synced_on,
            "importFile": os.path.basename(args.input),
        },
        "league": parse_league(workbook),
        "history": parse_history(workbook),
        "teams": parse_teams(workbook),
        "greats": parse_greats(workbook),
        "records": parse_records(workbook),
    }

    raw_json = json.dumps(payload, indent=2, ensure_ascii=False)
    with open(args.output, "w", encoding="utf8") as handle:
        handle.write(f"export const historySheet = {raw_json};\n")

    print(f"Imported history sheet from {args.input}")
    print(f"Wrote snapshot to {args.output}")
    print(f"syncedOn={synced_on}")


if __name__ == "__main__":
    main()
